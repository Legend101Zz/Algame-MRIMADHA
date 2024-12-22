"""
Results analysis panel implementation.

This panel provides:
1. Backtest results visualization
2. Performance metrics analysis
3. Trade list and details
4. Report generation

The panel is organized into tabs for different analysis views:
- Summary: Overview and key metrics
- Equity: Equity curve and drawdowns
- Trades: Detailed trade list and statistics
- Analysis: Advanced performance analysis
"""

import tkinter as tk
from tkinter import ttk, messagebox
from typing import Dict, Optional, Any, List
import pandas as pd
import numpy as np
from datetime import datetime
import logging

from ...core.engine import BacktestResult
from ..components.chart import Chart

logger = logging.getLogger(__name__)

class ResultsPanel(ttk.Frame):
    """
    Backtest results analysis panel.

    Features:
    - Multiple result views
    - Interactive charts
    - Trade analysis
    - Report export

    The panel uses a notebook layout with tabs for different analysis views.
    Each view is updated when new results are loaded.
    """

    def __init__(self, master):
        """Initialize panel."""
        super().__init__(master)

        # State
        self.results: Optional[BacktestResult] = None
        self._current_trade = None

        # Create widgets
        self._create_widgets()
        self._create_layout()
        self._bind_events()

        logger.debug("Initialized ResultsPanel")

    def _create_widgets(self):
        """Create panel widgets."""
        # Notebook for views
        self.notebook = ttk.Notebook(self)

        # Summary tab
        self.summary_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.summary_frame, text="Summary")
        self._create_summary_tab()

        # Equity tab
        self.equity_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.equity_frame, text="Equity")
        self._create_equity_tab()

        # Trades tab
        self.trades_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.trades_frame, text="Trades")
        self._create_trades_tab()

        # Analysis tab
        self.analysis_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.analysis_frame, text="Analysis")
        self._create_analysis_tab()

        # Toolbar
        self.toolbar = ttk.Frame(self)
        ttk.Button(
            self.toolbar,
            text="Export Results",
            command=self._export_results
        ).pack(side='left', padx=2)

        ttk.Button(
            self.toolbar,
            text="Generate Report",
            command=self._generate_report
        ).pack(side='left', padx=2)

    def _create_summary_tab(self):
        """Create summary view."""
        # Metrics frame
        metrics_frame = ttk.LabelFrame(self.summary_frame, text="Key Metrics")
        metrics_frame.pack(fill='x', padx=5, pady=5)

        # Metric grid
        self.metric_labels = {}
        metrics = [
            ('Total Return', 'total_return', '{:.2f}%'),
            ('Sharpe Ratio', 'sharpe_ratio', '{:.2f}'),
            ('Max Drawdown', 'max_drawdown', '{:.2f}%'),
            ('Win Rate', 'win_rate', '{:.1f}%'),
            ('Profit Factor', 'profit_factor', '{:.2f}'),
            ('Total Trades', 'total_trades', '{:d}')
        ]

        for i, (label, key, fmt) in enumerate(metrics):
            row = i // 3
            col = i % 3

            # Label
            ttk.Label(
                metrics_frame,
                text=f"{label}:"
            ).grid(row=row, column=col*2, padx=5, pady=2, sticky='e')

            # Value
            value_label = ttk.Label(metrics_frame, text="-")
            value_label.grid(row=row, column=col*2+1, padx=5, pady=2, sticky='w')
            self.metric_labels[key] = (value_label, fmt)

        # Trade summary
        summary_frame = ttk.LabelFrame(self.summary_frame, text="Trade Summary")
        summary_frame.pack(fill='both', expand=True, padx=5, pady=5)

        self.summary_text = tk.Text(
            summary_frame,
            height=10,
            wrap='word',
            state='disabled'
        )
        self.summary_text.pack(fill='both', expand=True, padx=5, pady=5)

    def _create_equity_tab(self):
        """Create equity curve view."""
        # Chart
        self.equity_chart = Chart(self.equity_frame)
        self.equity_chart.pack(fill='both', expand=True, padx=5, pady=5)

        # Controls
        controls = ttk.Frame(self.equity_frame)
        controls.pack(fill='x', padx=5, pady=5)

        # Plot type
        ttk.Label(controls, text="Plot:").pack(side='left')
        self.plot_type = tk.StringVar(value='equity')
        ttk.Combobox(
            controls,
            textvariable=self.plot_type,
            values=['equity', 'drawdown', 'returns'],
            state='readonly',
            width=10
        ).pack(side='left', padx=5)

        # Display options
        self.show_trades = tk.BooleanVar(value=True)
        ttk.Checkbutton(
            controls,
            text="Show Trades",
            variable=self.show_trades
        ).pack(side='left', padx=10)

    def _create_trades_tab(self):
        """Create trades list view."""
        # Trade list
        list_frame = ttk.Frame(self.trades_frame)
        list_frame.pack(fill='both', expand=True, padx=5, pady=5)

        self.trade_list = ttk.Treeview(
            list_frame,
            columns=('entry', 'exit', 'type', 'size', 'pnl', 'return'),
            show='headings',
            selectmode='browse'
        )

        self.trade_list.heading('entry', text='Entry')
        self.trade_list.heading('exit', text='Exit')
        self.trade_list.heading('type', text='Type')
        self.trade_list.heading('size', text='Size')
        self.trade_list.heading('pnl', text='P&L')
        self.trade_list.heading('return', text='Return %')

        # Scrollbar
        scrollbar = ttk.Scrollbar(
            list_frame,
            orient='vertical',
            command=self.trade_list.yview
        )
        self.trade_list.configure(yscrollcommand=scrollbar.set)

        # Layout
        self.trade_list.pack(side='left', fill='both', expand=True)
        scrollbar.pack(side='right', fill='y')

        # Trade details
        details_frame = ttk.LabelFrame(self.trades_frame, text="Trade Details")
        details_frame.pack(fill='x', padx=5, pady=5)

        self.trade_details = tk.Text(
            details_frame,
            height=5,
            wrap='word',
            state='disabled'
        )
        self.trade_details.pack(fill='x', padx=5, pady=5)

    def _create_analysis_tab(self):
        """Create detailed analysis view."""
        # Metrics selection
        metrics_frame = ttk.LabelFrame(self.analysis_frame, text="Analysis")
        metrics_frame.pack(fill='x', padx=5, pady=5)

        ttk.Label(metrics_frame, text="Metric:").pack(side='left', padx=5)
        self.analysis_metric = tk.StringVar(value='returns')
        ttk.Combobox(
            metrics_frame,
            textvariable=self.analysis_metric,
            values=[
                'returns',
                'drawdowns',
                'trade_pnl',
                'win_rate',
                'position_size'
            ],
            state='readonly',
            width=15
        ).pack(side='left', padx=5)

        # Analysis chart
        self.analysis_chart = Chart(self.analysis_frame)
        self.analysis_chart.pack(fill='both', expand=True, padx=5, pady=5)

    def _create_layout(self):
        """Create panel layout."""
        # Notebook
        self.notebook.pack(fill='both', expand=True)

        # Toolbar
        self.toolbar.pack(fill='x', padx=5, pady=5)

    def _bind_events(self):
        """Bind event handlers."""
        self.plot_type.trace('w', lambda *args: self._update_equity_chart())
        self.show_trades.trace('w', lambda *args: self._update_equity_chart())
        self.analysis_metric.trace('w', lambda *args: self._update_analysis())
        self.trade_list.bind('<<TreeviewSelect>>', self._on_trade_select)

    def show_results(self, results: BacktestResult):
        """Show backtest results."""
        self.results = results

        # Update all views
        self._update_summary()
        self._update_equity_chart()
        self._update_trades()
        self._update_analysis()

        # Switch to summary tab
        self.notebook.select(0)

        logger.info("Updated results display")

    def _update_summary(self):
        """Update summary view."""
        if not self.results:
            return

        # Update metrics
        for key, (label, fmt) in self.metric_labels.items():
            value = self.results.metrics.get(key, 0)
            label['text'] = fmt.format(value)

        # Update trade summary
        self.summary_text['state'] = 'normal'
        self.summary_text.delete('1.0', 'end')

        summary = self._generate_summary()
        self.summary_text.insert('1.0', summary)
        self.summary_text['state'] = 'disabled'

    def _generate_summary(self) -> str:
        """Generate summary text."""
        metrics = self.results.metrics
        trades = self.results.trades

        long_trades = len([t for t in trades if t.type == 'long'])
        short_trades = len([t for t in trades if t.type == 'short'])

        summary = [
            f"Backtest Period: {self.results.start_date:%Y-%m-%d} to {self.results.end_date:%Y-%m-%d}",
            f"Initial Capital: ${self.results.config.initial_capital:,.2f}",
            "",
            f"Total Trades: {len(trades)}",
            f"Long Trades: {long_trades}",
            f"Short Trades: {short_trades}",
            "",
            f"Net Profit: ${metrics.get('net_profit', 0):,.2f}",
            f"Profit Factor: {metrics.get('profit_factor', 0):.2f}",
            f"Win Rate: {metrics.get('win_rate', 0):.1f}%",
            "",
            f"Max Drawdown: {metrics.get('max_drawdown', 0):.2f}%",
            f"Avg. Drawdown: {metrics.get('avg_drawdown', 0):.2f}%",
            f"Recovery Factor: {metrics.get('recovery_factor', 0):.2f}",
            "",
            f"Sharpe Ratio: {metrics.get('sharpe_ratio', 0):.2f}",
            f"Sortino Ratio: {metrics.get('sortino_ratio', 0):.2f}",
            f"Calmar Ratio: {metrics.get('calmar_ratio', 0):.2f}"
        ]

        return "\n".join(summary)

    def _update_equity_chart(self):
        """Update equity chart."""
        if not self.results:
            return

        plot_type = self.plot_type.get()

        if plot_type == 'equity':
            data = self.results.equity_curve
            title = "Equity Curve"
            y_label = "Equity ($)"
        elif plot_type == 'drawdown':
            data = self.results.drawdowns
            title = "Drawdown"
            y_label = "Drawdown (%)"
        else:  # returns
            data = self.results.returns
            title = "Returns"
            y_label = "Return (%)"

        # Plot data
        self.equity_chart.plot(
            data,
            title=title,
            y_label=y_label
        )

        # Add trade markers if enabled
        if self.show_trades.get():
            self._add_trade_markers()

    def _add_trade_markers(self):
        """Add trade markers to equity chart."""
        for trade in self.results.trades:
            # Plot entry
            self.equity_chart.add_marker(
                trade.entry_time,
                self.results.equity_curve[trade.entry_time],
                'entry',
                color='green' if trade.type == 'long' else 'red'
            )

            # Plot exit
            if trade.exit_time:
                self.equity_chart.add_marker(
                    trade.exit_time,
                    self.results.equity_curve[trade.exit_time],
                    'exit',
                    color='red' if trade.type == 'long' else 'green'
                )

    def _update_trades(self):
        """Update trades list."""
        # Clear list
        for item in self.trade_list.get_children():
            self.trade_list.delete(item)

        if not self.results:
            return

        # Add trades
        for trade in self.results.trades:
            self.trade_list.insert('', 'end', values=(
                trade.entry_time.strftime('%Y-%m-%d %H:%M'),
                trade.exit_time.strftime('%Y-%m-%d %H:%M') if trade.exit_time else '-',
                trade.type,
                f"{trade.size:.2f}",
                f"${trade.pnl:.2f}",
                f"{trade.return_pct:.1f}%"
            ))

    def _on_trade_select(self, event):
        """Handle trade selection."""
        selection = self.trade_list.selection()
        if not selection:
            return

        # Get selected trade
        index = self.trade_list.index(selection[0])
        trade = self.results.trades[index]
        self._current_trade = trade

        # Update details
        self.trade_details['state'] = 'normal'
        self.trade_details.delete('1.0', 'end')

        # Format trade details
        details = [
            f"Entry: {trade.entry_time:%Y-%m-%d %H:%M} @ ${trade.entry_price:.2f}",
            f"Exit: {trade.exit_time:%Y-%m-%d %H:%M} @ ${trade.exit_price:.2f}" if trade.exit_time else "Exit: Still Open",
            f"Type: {trade.type.title()}",
            f"Size: {trade.size:.2f}",
            f"P&L: ${trade.pnl:.2f} ({trade.return_pct:.1f}%)",
            f"Duration: {trade.duration:.1f} days" if trade.duration else "Duration: N/A",
            f"Fees: ${trade.fees:.2f}"
        ]

        self.trade_details.insert('1.0', "\n".join(details))
        self.trade_details['state'] = 'disabled'

    def _update_analysis(self):
        """Update analysis view."""
        if not self.results:
            return

        metric = self.analysis_metric.get()

        if metric == 'returns':
            self._plot_return_distribution()
        elif metric == 'drawdowns':
            self._plot_drawdown_analysis()
        elif metric == 'trade_pnl':
            self._plot_trade_analysis()
        elif metric == 'win_rate':
            self._plot_win_rate_analysis()
        elif metric == 'position_size':
            self._plot_position_analysis()

    def _plot_return_distribution(self):
        """Plot return distribution analysis."""
        returns = self.results.returns.dropna()

        # Calculate statistics
        mean = returns.mean()
        std = returns.std()
        skew = returns.skew()
        kurtosis = returns.kurtosis()

        # Plot histogram
        self.analysis_chart.plot_histogram(
            returns,
            title="Return Distribution",
            xlabel="Return (%)",
            ylabel="Frequency",
            stats={
                "Mean": f"{mean:.2f}%",
                "Std Dev": f"{std:.2f}%",
                "Skewness": f"{skew:.2f}",
                "Kurtosis": f"{kurtosis:.2f}"
            }
        )

    def _plot_drawdown_analysis(self):
        """Plot drawdown analysis."""
        drawdowns = self.results.drawdowns

        # Calculate statistics
        max_dd = drawdowns.min()
        avg_dd = drawdowns.mean()
        dd_std = drawdowns.std()

        # Plot drawdown distribution
        self.analysis_chart.plot_histogram(
            drawdowns,
            title="Drawdown Analysis",
            xlabel="Drawdown (%)",
            ylabel="Frequency",
            stats={
                "Max": f"{max_dd:.2f}%",
                "Average": f"{avg_dd:.2f}%",
                "Std Dev": f"{dd_std:.2f}%"
            }
        )

    def _plot_trade_analysis(self):
        """Plot trade P&L analysis."""
        pnls = [t.pnl for t in self.results.trades]

        # Calculate statistics
        avg_pnl = np.mean(pnls)
        best_trade = max(pnls)
        worst_trade = min(pnls)

        # Plot P&L distribution
        self.analysis_chart.plot_histogram(
            pnls,
            title="Trade P&L Distribution",
            xlabel="P&L ($)",
            ylabel="Frequency",
            stats={
                "Average": f"${avg_pnl:.2f}",
                "Best": f"${best_trade:.2f}",
                "Worst": f"${worst_trade:.2f}"
            }
        )

    def _plot_win_rate_analysis(self):
        """Plot win rate analysis."""
        trades = self.results.trades

        # Calculate win rates by:
        # - Trade type (long/short)
        # - Time of day
        # - Day of week
        # - Month

        # By type
        long_trades = [t for t in trades if t.type == 'long']
        short_trades = [t for t in trades if t.type == 'short']

        long_wins = len([t for t in long_trades if t.pnl > 0])
        short_wins = len([t for t in short_trades if t.pnl > 0])

        win_rates = {
            'Long': long_wins / len(long_trades) if long_trades else 0,
            'Short': short_wins / len(short_trades) if short_trades else 0
        }

        # Plot win rates
        self.analysis_chart.plot_bar(
            list(win_rates.keys()),
            list(win_rates.values()),
            title="Win Rate Analysis",
            xlabel="Trade Type",
            ylabel="Win Rate (%)",
            percentage=True
        )

    def _plot_position_analysis(self):
        """Plot position size analysis."""
        sizes = [t.size for t in self.results.trades]

        # Calculate statistics
        avg_size = np.mean(sizes)
        max_size = max(sizes)
        min_size = min(sizes)

        # Plot size distribution
        self.analysis_chart.plot_histogram(
            sizes,
            title="Position Size Distribution",
            xlabel="Size",
            ylabel="Frequency",
            stats={
                "Average": f"{avg_size:.2f}",
                "Max": f"{max_size:.2f}",
                "Min": f"{min_size:.2f}"
            }
        )

    def _export_results(self):
        """Export results to file."""
        if not self.results:
            messagebox.showwarning(
                "No Results",
                "No results available to export."
            )
            return

        from tkinter import filedialog

        # Get output file
        filename = filedialog.asksaveasfilename(
            defaultextension=".csv",
            filetypes=[
                ("CSV files", "*.csv"),
                ("Excel files", "*.xlsx"),
                ("JSON files", "*.json")
            ]
        )

        if not filename:
            return

        try:
            # Export based on file type
            ext = filename.split('.')[-1].lower()

            if ext == 'csv':
                self._export_csv(filename)
            elif ext == 'xlsx':
                self._export_excel(filename)
            else:
                self._export_json(filename)

            messagebox.showinfo(
                "Export Complete",
                f"Results exported to {filename}"
            )

        except Exception as e:
            messagebox.showerror(
                "Export Error",
                f"Failed to export results: {str(e)}"
            )

    def _export_csv(self, filename: str):
        """Export results to CSV."""
        # Export trade list
        trades_df = pd.DataFrame([
            {
                'Entry Time': t.entry_time,
                'Exit Time': t.exit_time,
                'Type': t.type,
                'Size': t.size,
                'Entry Price': t.entry_price,
                'Exit Price': t.exit_price,
                'P&L': t.pnl,
                'Return %': t.return_pct,
                'Fees': t.fees
            }
            for t in self.results.trades
        ])

        trades_df.to_csv(filename, index=False)

    def _export_excel(self, filename: str):
        """Export results to Excel."""
        from openpyxl import Workbook
        from openpyxl.utils.dataframe import dataframe_to_rows

        wb = Workbook()

        # Summary sheet
        ws = wb.active
        ws.title = "Summary"

        # Add metrics
        ws['A1'] = "Metric"
        ws['B1'] = "Value"

        row = 2
        for key, value in self.results.metrics.items():
            ws[f'A{row}'] = key
            ws[f'B{row}'] = value
            row += 1

        # Trade list sheet
        ws = wb.create_sheet("Trades")

        trades_df = pd.DataFrame([
            {
                'Entry Time': t.entry_time,
                'Exit Time': t.exit_time,
                'Type': t.type,
                'Size': t.size,
                'Entry Price': t.entry_price,
                'Exit Price': t.exit_price,
                'P&L': t.pnl,
                'Return %': t.return_pct,
                'Fees': t.fees
            }
            for t in self.results.trades
        ])

        for r in dataframe_to_rows(trades_df, index=False, header=True):
            ws.append(r)

        # Equity curve sheet
        ws = wb.create_sheet("Equity")

        equity_df = pd.DataFrame({
            'Equity': self.results.equity_curve,
            'Drawdown': self.results.drawdowns,
            'Returns': self.results.returns
        })

        for r in dataframe_to_rows(equity_df, index=True, header=True):
            ws.append(r)

        wb.save(filename)

    def _export_json(self, filename: str):
        """Export results to JSON."""
        import json

        data = {
            'metadata': {
                'start_date': self.results.start_date.isoformat(),
                'end_date': self.results.end_date.isoformat(),
                'config': vars(self.results.config)
            },
            'metrics': self.results.metrics,
            'trades': [
                {
                    'entry_time': t.entry_time.isoformat(),
                    'exit_time': t.exit_time.isoformat() if t.exit_time else None,
                    'type': t.type,
                    'size': t.size,
                    'entry_price': t.entry_price,
                    'exit_price': t.exit_price,
                    'pnl': t.pnl,
                    'return_pct': t.return_pct,
                    'fees': t.fees
                }
                for t in self.results.trades
            ],
            'equity_curve': self.results.equity_curve.to_dict(),
            'drawdowns': self.results.drawdowns.to_dict(),
            'returns': self.results.returns.to_dict()
        }

        with open(filename, 'w') as f:
            json.dump(data, f, indent=4)

    def _generate_report(self):
        """Generate PDF report."""
        messagebox.showinfo(
            "Not Implemented",
            "Report generation coming soon!"
        )

    def reset(self):
        """Reset panel state."""
        self.results = None
        self._current_trade = None

        # Clear displays
        self._update_summary()
        self._update_trades()
        self._update_analysis()
